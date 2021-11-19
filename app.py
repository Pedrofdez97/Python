from tkinter import *
from html.parser import HTMLParser
import bs4
import openpyxl as openpyxl
import requests
from bs4 import BeautifulSoup



#Diccionario que creamos para guardar los datos de las bicicletas.
bicicletas = {
    "url_imagen": "",
    "marca": "",
    "precio_inicial": "",
    "precio_final": "",
    "descuento": ""
    }
# Dirección de la página
url = "https://www.bikester.es/bicicletas/?page="

#Función para cargar las paginas.
def cargar_paginas(link):
    #Página con la que queremos trabajar.
    html = requests.get(link).content
    pagina = BeautifulSoup(html, "html.parser")
    return pagina

#Función para insertar los elementos de la página.
def cargar_elementos(paginas):
    listabicicletas = list()
    lista_elementos = paginas.findAll("div", {"class": "product-tile-inner"})
    a = 0

    #Busqueda de los elementos que queremos de la página.
    for bicis in lista_elementos:
        url_imagen = bicis.find("div",{"class": "product-image"}).find("img").attrs["src"]
        marca = bicis.find("div", {"class": "cyc-typo_subheader cyc-color-text"}).text.replace("\n", "").replace("  ","")
        precio_inicial = bicis.find("div", {"class":"product-price"}).find("span", {"class": "price-standard"}).text.replace("€","").replace("\n","").replace(" ","")
        descuento_comprobacion = bicis.find("div", {"class": "discount cyc-padding_leftright-1 cyc-typo_body"})
        precio_final_comprobacion = bicis.find("div", {"class": "product-price"}).find("span", {"class": "price-sales"})

    # Comprobaciones de si tiene descuento de las bicicletas.
        if descuento_comprobacion is None:
            descuento = "sin descuento"
        else:
            descuento = bicis.find("div", {"class": "discount cyc-padding_leftright-1 cyc-typo_body"}).text.replace("€","").replace("\n","").replace(" ","").replace("-","").replace("%","")

    # Comprobaciones del precio final de las bicicletas si no tiene que sea igual que el precio inicial.
        if precio_final_comprobacion is None:
            precio_final = precio_inicial
        else:
            precio_final = bicis.find("div", {"class": "product-price"}).find("span",{"class": "price-sales"}).text.replace("€", "").replace("\n", "").replace(" ", "")


        #Guardar datos en diccionario
        tabla_bicicletas = bicicletas.copy()
        bicicletas["url_imagen"] = url_imagen
        bicicletas["marca"] = marca
        bicicletas["precio_inicial"] = precio_inicial
        bicicletas["precio_final"] = precio_final
        bicicletas["descuento"] = descuento
        a += 1
        if a > 1:
            listabicicletas.append(tabla_bicicletas)


    return listabicicletas

#Función para cargar las páginas que queremos de enlace.
def cargar_todas_las_paginas():
    lista_paginas = list()
    for num_pagina in range(5):
        pagina = cargar_paginas(url+str(num_pagina))
        lista_paginas.append(pagina)
    return lista_paginas

#Función para meter todos los elementos en una lista nueva.
def cargar_los_elementos():
    lista_bici_completa = list()
    lista_paginas = cargar_todas_las_paginas()

    for paginas in lista_paginas:
        lista_bici_paginas = cargar_elementos(paginas)
        lista_bici_completa.extend(lista_bici_paginas)
    return lista_bici_completa

#Función para crear el excel y meter la lista que hemos guardado anteriormente.
def excel():
    list_elementos = cargar_los_elementos()

    documento_excel = openpyxl.Workbook()

    hoja1 = documento_excel.active

    hoja1.title = "Bicicletas"

    count = 1

    for key in bicicletas.keys():
        hoja1.cell(row=1, column=count, value=key)
        count += 1

    fila = 2

    for bici in list_elementos:
        for columna in range(len(bicicletas)):
            hoja1.cell(row=fila, column=columna + 1, value=bici[list(bici.keys())[columna]])
        fila += 1

    documento_excel.save("bicis.xlsx")
